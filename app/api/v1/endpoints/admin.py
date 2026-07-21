from __future__ import annotations

from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from io import BytesIO

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import func, or_, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, selectinload

from app.api.deps import require_admin
from app.core.categories import WOMEN_CLUB_CATEGORY_SLUGS
from app.core.config import settings
from app.core.security import hash_password
from app.db.session import get_db
from app.models.category import Category
from app.models.city import City
from app.models.client import ClientProfile
from app.models.giveaway import Giveaway, GiveawayNumber, GiveawayPrize
from app.models.engagement import (
    BloomDailyTask,
    BloomGardenSettings,
    BloomSpecialAnswer,
    BloomSpecialOption,
    BloomSpecialQuestion,
    BloomSpecialSubmission,
    BloomSpecialTask,
    PartnerBotAccess,
)
from app.models.lead import LeadClick
from app.models.landing import LandingSettings
from app.models.partner import Partner, PartnerOffer, PartnerPhoto, PartnerQrLink
from app.models.payment import PaymentRequest, PaymentRequestStatus, Subscription, SubscriptionStatus
from app.models.user import AdminUser, User, UserRole
from app.models.verification import PrivilegeVerificationSession
from app.schemas.activity import ActivityFeedRead
from app.schemas.admin import (
    AdminManagedUserCreate,
    AdminManagedUserRead,
    AdminManagedUserUpdate,
    AdminDeleteUserResponse,
    AdminVerificationRead,
    ContentReviewOfferRead,
    ContentReviewPhotoRead,
    ContentReviewRead,
    CategoryCreate,
    CategoryRead,
    CategoryUpdate,
    CityCreate,
    CityRead,
    CityUpdate,
    LeadStatsRead,
    PartnerCreate,
    PartnerOfferCreate,
    PartnerOfferRead,
    PartnerOfferUpdate,
    PartnerPhotoRead,
    PartnerPhotoUpdate,
    PartnerPhotoUploadResponse,
    PartnerQrLinkCreate,
    PartnerQrLinkRead,
    PartnerQrLinkUpdate,
    PartnerRead,
    PartnerUpdate,
)
from app.schemas.auth import AdminUserRead
from app.schemas.giveaway import GiveawayRead, GiveawayWrite, GiveawayPrizeRead, GiveawayPrizeWrite
from app.schemas.engagement import (
    BloomTaskPatch,
    BloomTaskRead,
    BloomTaskWrite,
    BloomGardenSettingsPatch,
    BloomGardenSettingsRead,
    BloomSpecialAnalyticsRead,
    BloomSpecialOptionAnalyticsRead,
    BloomSpecialQuestionAnalyticsRead,
    BloomSpecialQuestionWrite,
    BloomSpecialSubmissionRead,
    BloomSpecialTaskPatch,
    BloomSpecialTaskRead,
    BloomSpecialTaskWrite,
    FlowerLeaderboardRewardRead,
    FlowerLeaderboardSettleRequest,
    PartnerBotAccessPatch,
    PartnerBotAccessRead,
    PartnerBotAccessWrite,
)
from app.schemas.landing import LandingSettingsRead, LandingSettingsUpdate
from app.schemas.partner import PartnerAnalyticsRead
from app.schemas.payment import AdminPaymentRequestRead, PaymentRequestApprove, PaymentRequestReject
from app.services.activity_feed import build_admin_activity_feed
from app.services.admin_user_delete_service import delete_user_with_relations
from app.services.landing_settings import build_admin_landing_settings_read, get_or_create_landing_settings, normalize_giveaway_items
from app.services.image_uploads import save_partner_image_upload, save_partner_offer_image_upload, save_partner_photo_image_upload, validate_image_kind
from app.services.partner_analytics import build_partner_analytics
from app.services.privilege_verifications import (
    apply_verification_status_filter,
    as_aware_utc,
    normalize_expired_verifications,
    ttl_seconds,
)
from app.services.social_subscriptions import recheck_giveaway_social_subscriptions, is_number_active
from app.services.engagement import club_today, garden_settings, settle_flower_leaderboard
from app.services.qr_links import (
    generate_qr_slug,
    is_valid_qr_slug,
    normalize_qr_slug,
    qr_link_to_read,
)

router = APIRouter(prefix="/admin", tags=["admin"])


def _partner_access_read(access: PartnerBotAccess) -> PartnerBotAccessRead:
    return PartnerBotAccessRead(
        id=access.id,
        partner_id=access.partner_id,
        partner_name=access.partner.name,
        provider=access.provider,
        provider_user_id=access.provider_user_id,
        username=access.username,
        display_name=access.display_name,
        is_active=access.is_active,
        activation_count=access.activation_count,
        last_activity_at=access.last_activity_at,
        created_at=access.created_at,
    )


@router.get("/partner-accesses", response_model=list[PartnerBotAccessRead])
def list_partner_accesses(admin: AdminUser = Depends(require_admin), db: Session = Depends(get_db)) -> list[PartnerBotAccessRead]:
    _ = admin
    rows = db.execute(select(PartnerBotAccess).options(selectinload(PartnerBotAccess.partner)).order_by(PartnerBotAccess.created_at.desc())).scalars().all()
    return [_partner_access_read(row) for row in rows]


@router.post("/partner-accesses", response_model=PartnerBotAccessRead, status_code=status.HTTP_201_CREATED)
def create_partner_access(payload: PartnerBotAccessWrite, admin: AdminUser = Depends(require_admin), db: Session = Depends(get_db)) -> PartnerBotAccessRead:
    _ = admin
    partner = db.get(Partner, payload.partner_id)
    if partner is None:
        raise HTTPException(status_code=404, detail="Partner not found")
    access = PartnerBotAccess(**payload.model_dump())
    access.provider_user_id = access.provider_user_id.strip()
    access.display_name = access.display_name.strip()
    access.username = (access.username or "").strip() or None
    db.add(access)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=409, detail="This bot account already has partner access") from None
    db.refresh(access)
    access.partner = partner
    return _partner_access_read(access)


@router.patch("/partner-accesses/{access_id}", response_model=PartnerBotAccessRead)
def update_partner_access(access_id: int, payload: PartnerBotAccessPatch, admin: AdminUser = Depends(require_admin), db: Session = Depends(get_db)) -> PartnerBotAccessRead:
    _ = admin
    access = db.get(PartnerBotAccess, access_id)
    if access is None:
        raise HTTPException(status_code=404, detail="Partner access not found")
    for field, value in payload.model_dump(exclude_unset=True).items():
        if field == "partner_id" and db.get(Partner, value) is None:
            raise HTTPException(status_code=404, detail="Partner not found")
        setattr(access, field, value.strip() if isinstance(value, str) else value)
    db.commit()
    access = db.execute(select(PartnerBotAccess).options(selectinload(PartnerBotAccess.partner)).where(PartnerBotAccess.id == access_id)).scalar_one()
    return _partner_access_read(access)


@router.get("/flower/tasks", response_model=list[BloomTaskRead])
def list_flower_tasks(admin: AdminUser = Depends(require_admin), db: Session = Depends(get_db)) -> list[BloomTaskRead]:
    _ = admin
    return list(db.execute(select(BloomDailyTask).order_by(BloomDailyTask.sort_order, BloomDailyTask.id)).scalars().all())


@router.post("/flower/tasks", response_model=BloomTaskRead, status_code=status.HTTP_201_CREATED)
def create_flower_task(payload: BloomTaskWrite, admin: AdminUser = Depends(require_admin), db: Session = Depends(get_db)) -> BloomTaskRead:
    _ = admin
    task = BloomDailyTask(**payload.model_dump())
    db.add(task)
    db.commit()
    db.refresh(task)
    return task


@router.patch("/flower/tasks/{task_id}", response_model=BloomTaskRead)
def update_flower_task(task_id: int, payload: BloomTaskPatch, admin: AdminUser = Depends(require_admin), db: Session = Depends(get_db)) -> BloomTaskRead:
    _ = admin
    task = db.get(BloomDailyTask, task_id)
    if task is None:
        raise HTTPException(status_code=404, detail="Flower task not found")
    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(task, field, value)
    db.commit()
    db.refresh(task)
    return task


@router.get("/flower/settings", response_model=BloomGardenSettingsRead)
def read_flower_settings(admin: AdminUser = Depends(require_admin), db: Session = Depends(get_db)) -> BloomGardenSettingsRead:
    _ = admin
    settings_row = garden_settings(db)
    if settings_row not in db:
        db.add(settings_row)
        db.commit()
        db.refresh(settings_row)
    return BloomGardenSettingsRead.model_validate(settings_row, from_attributes=True)


@router.patch("/flower/settings", response_model=BloomGardenSettingsRead)
def update_flower_settings(payload: BloomGardenSettingsPatch, admin: AdminUser = Depends(require_admin), db: Session = Depends(get_db)) -> BloomGardenSettingsRead:
    _ = admin
    settings_row = garden_settings(db)
    if settings_row not in db:
        db.add(settings_row)
    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(settings_row, field, value)
    db.commit()
    db.refresh(settings_row)
    return BloomGardenSettingsRead.model_validate(settings_row, from_attributes=True)


def _special_task_read(db: Session, task: BloomSpecialTask) -> BloomSpecialTaskRead:
    count = db.execute(select(func.count(BloomSpecialSubmission.id)).where(BloomSpecialSubmission.task_id == task.id)).scalar_one()
    result = BloomSpecialTaskRead.model_validate(task, from_attributes=True)
    return result.model_copy(update={"submissions_count": int(count or 0)})


def _ensure_no_special_task_overlap(db: Session, starts_on: date, ends_on: date, *, exclude_id: int | None = None) -> None:
    query = select(BloomSpecialTask.id).where(
        BloomSpecialTask.is_active.is_(True),
        BloomSpecialTask.starts_on <= ends_on,
        BloomSpecialTask.ends_on >= starts_on,
    )
    if exclude_id is not None:
        query = query.where(BloomSpecialTask.id != exclude_id)
    if db.execute(query).scalars().first() is not None:
        raise HTTPException(status_code=409, detail="Another active special task overlaps this period")


@router.get("/flower/special-tasks", response_model=list[BloomSpecialTaskRead])
def list_special_tasks(admin: AdminUser = Depends(require_admin), db: Session = Depends(get_db)) -> list[BloomSpecialTaskRead]:
    _ = admin
    tasks = db.execute(
        select(BloomSpecialTask)
        .options(selectinload(BloomSpecialTask.questions).selectinload(BloomSpecialQuestion.options))
        .order_by(BloomSpecialTask.starts_on.desc(), BloomSpecialTask.id.desc())
    ).scalars().all()
    return [_special_task_read(db, task) for task in tasks]


@router.post("/flower/special-tasks", response_model=BloomSpecialTaskRead, status_code=status.HTTP_201_CREATED)
def create_special_task(payload: BloomSpecialTaskWrite, admin: AdminUser = Depends(require_admin), db: Session = Depends(get_db)) -> BloomSpecialTaskRead:
    _ = admin
    if payload.ends_on < payload.starts_on:
        raise HTTPException(status_code=422, detail="ends_on must not be before starts_on")
    if payload.is_active:
        _ensure_no_special_task_overlap(db, payload.starts_on, payload.ends_on)
    task = BloomSpecialTask(**payload.model_dump())
    db.add(task)
    db.commit()
    db.refresh(task)
    return _special_task_read(db, task)


@router.patch("/flower/special-tasks/{task_id}", response_model=BloomSpecialTaskRead)
def update_special_task(task_id: int, payload: BloomSpecialTaskPatch, admin: AdminUser = Depends(require_admin), db: Session = Depends(get_db)) -> BloomSpecialTaskRead:
    _ = admin
    task = db.execute(select(BloomSpecialTask).options(selectinload(BloomSpecialTask.questions).selectinload(BloomSpecialQuestion.options)).where(BloomSpecialTask.id == task_id)).scalar_one_or_none()
    if task is None:
        raise HTTPException(status_code=404, detail="Special task not found")
    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(task, field, value)
    if task.ends_on < task.starts_on:
        raise HTTPException(status_code=422, detail="ends_on must not be before starts_on")
    if task.is_active:
        _ensure_no_special_task_overlap(db, task.starts_on, task.ends_on, exclude_id=task.id)
    db.commit()
    db.refresh(task)
    return _special_task_read(db, task)


@router.post("/flower/special-tasks/{task_id}/questions", response_model=BloomSpecialTaskRead, status_code=status.HTTP_201_CREATED)
def add_special_question(task_id: int, payload: BloomSpecialQuestionWrite, admin: AdminUser = Depends(require_admin), db: Session = Depends(get_db)) -> BloomSpecialTaskRead:
    _ = admin
    task = db.get(BloomSpecialTask, task_id)
    if task is None:
        raise HTTPException(status_code=404, detail="Special task not found")
    sort_order = int(db.execute(select(func.count(BloomSpecialQuestion.id)).where(BloomSpecialQuestion.task_id == task_id)).scalar_one() or 0)
    question = BloomSpecialQuestion(task_id=task_id, prompt=payload.prompt.strip(), sort_order=sort_order)
    db.add(question)
    db.flush()
    for index, label in enumerate(payload.options):
        db.add(BloomSpecialOption(question_id=question.id, label=label, sort_order=index))
    db.commit()
    task = db.execute(select(BloomSpecialTask).options(selectinload(BloomSpecialTask.questions).selectinload(BloomSpecialQuestion.options)).where(BloomSpecialTask.id == task_id)).scalar_one()
    return _special_task_read(db, task)


@router.get("/flower/special-tasks/{task_id}/analytics", response_model=BloomSpecialAnalyticsRead)
def special_task_analytics(task_id: int, admin: AdminUser = Depends(require_admin), db: Session = Depends(get_db)) -> BloomSpecialAnalyticsRead:
    _ = admin
    task = db.execute(select(BloomSpecialTask).options(selectinload(BloomSpecialTask.questions).selectinload(BloomSpecialQuestion.options)).where(BloomSpecialTask.id == task_id)).scalar_one_or_none()
    if task is None:
        raise HTTPException(status_code=404, detail="Special task not found")
    submissions = db.execute(
        select(BloomSpecialSubmission)
        .options(selectinload(BloomSpecialSubmission.answers))
        .where(BloomSpecialSubmission.task_id == task_id)
        .order_by(BloomSpecialSubmission.completed_at.desc())
    ).scalars().all()
    option_by_id = {option.id: option for question in task.questions for option in question.options}
    answer_counts: dict[int, int] = {}
    for submission in submissions:
        for answer in submission.answers:
            answer_counts[answer.option_id] = answer_counts.get(answer.option_id, 0) + 1
    question_stats = []
    for question in task.questions:
        total = sum(answer_counts.get(option.id, 0) for option in question.options)
        question_stats.append(BloomSpecialQuestionAnalyticsRead(
            question_id=question.id,
            prompt=question.prompt,
            options=[BloomSpecialOptionAnalyticsRead(option_id=option.id, label=option.label, count=answer_counts.get(option.id, 0), percent=round(answer_counts.get(option.id, 0) * 100 / total, 1) if total else 0.0) for option in question.options],
        ))
    profiles = {profile.id: profile for profile in db.execute(select(ClientProfile).options(selectinload(ClientProfile.user)).where(ClientProfile.id.in_([item.client_id for item in submissions]))).scalars().all()} if submissions else {}
    submission_rows = []
    for submission in submissions:
        profile = profiles.get(submission.client_id)
        submission_rows.append(BloomSpecialSubmissionRead(
            client_id=submission.client_id,
            full_name=profile.full_name if profile else None,
            email=profile.contact_email or profile.user.email if profile and profile.user else profile.contact_email if profile else None,
            phone=profile.user.phone if profile and profile.user else None,
            telegram_username=profile.telegram_username if profile else None,
            vk_username=profile.vk_username if profile else None,
            completed_at=submission.completed_at,
            answers=[option_by_id[answer.option_id].label for answer in submission.answers if answer.option_id in option_by_id],
        ))
    return BloomSpecialAnalyticsRead(task_id=task.id, title=task.title, submissions_count=len(submissions), questions=question_stats, submissions=submission_rows)


@router.post("/flower/settle", response_model=list[FlowerLeaderboardRewardRead])
def settle_flower(payload: FlowerLeaderboardSettleRequest, admin: AdminUser = Depends(require_admin), db: Session = Depends(get_db)) -> list[FlowerLeaderboardRewardRead]:
    _ = admin
    if payload.month.day != 1:
        raise HTTPException(status_code=422, detail="month must be the first day of a month")
    current_month = club_today().replace(day=1)
    if payload.month >= current_month:
        raise HTTPException(status_code=409, detail="The flower month has not finished yet")
    giveaway = db.get(Giveaway, payload.giveaway_id)
    if giveaway is None:
        raise HTTPException(status_code=404, detail="Giveaway not found")
    rewards = settle_flower_leaderboard(db, payload.month, giveaway)
    return [FlowerLeaderboardRewardRead(client_id=item.client_id, place=item.place, entries_count=item.entries_count) for item in rewards]

CITY_DUPLICATE_DETAIL = "City with this slug or name already exists"
CATEGORY_DUPLICATE_DETAIL = "Category with this slug already exists"
USER_DUPLICATE_DETAIL = "User with this email or phone already exists"
ALLOWED_USER_ROLES = tuple(role.value for role in UserRole)
PARTNER_TEXT_FIELDS = (
    "description",
    "address",
    "phone",
    "website_url",
    "social_url",
    "instagram_url",
    "vk_url",
    "telegram_url",
    "whatsapp_url",
    "map_url",
    "working_hours",
    "logo_url",
    "cover_url",
)
PARTNER_OFFER_TEXT_FIELDS = ("description", "benefit_text", "conditions", "image_url")


def require_legacy_content_write_enabled() -> None:
    if not settings.WEB_ADMIN_LEGACY_CONTENT_WRITE_ENABLED:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Legacy WEB content editing is disabled; use Content Admin API",
        )


def _giveaway_to_read(giveaway: Giveaway) -> GiveawayRead:
    return GiveawayRead(
        id=giveaway.id,
        title=giveaway.title,
        description=giveaway.description,
        is_active=giveaway.is_active,
        starts_at=giveaway.starts_at,
        ends_at=giveaway.ends_at,
        winners_count=giveaway.winners_count,
        created_at=giveaway.created_at,
        updated_at=giveaway.updated_at,
        prizes=[GiveawayPrizeRead(id=p.id, place_number=p.place_number, prize_title=p.prize_title, winner_provider=p.winner_provider, winner_provider_user_id=p.winner_provider_user_id, winning_number=p.winning_number) for p in sorted(giveaway.prizes, key=lambda item: item.place_number)],
        telegram_community_url=giveaway.telegram_community_url,
        telegram_chat_id=giveaway.telegram_chat_id,
        telegram_reward_enabled=giveaway.telegram_reward_enabled,
        telegram_reward_numbers=giveaway.telegram_reward_numbers,
        vk_community_url=giveaway.vk_community_url,
        vk_group_id=giveaway.vk_group_id,
        vk_reward_enabled=giveaway.vk_reward_enabled,
        vk_reward_numbers=giveaway.vk_reward_numbers,
    )

def _validate_giveaway_prizes(payload: GiveawayWrite) -> list[GiveawayPrizeWrite]:
    prizes = list(payload.prizes[: payload.winners_count])
    seen_places: set[int] = set()
    for prize in prizes:
        if prize.place_number < 1:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Prize place_number must be greater than or equal to 1")
        if prize.place_number in seen_places:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Prize place_number values must be unique")
        seen_places.add(prize.place_number)
    return prizes


def _copy_giveaway_prize_fields(prize: GiveawayPrize, payload: GiveawayPrizeWrite) -> None:
    prize.place_number = payload.place_number
    prize.prize_title = payload.prize_title.strip()
    prize.winner_provider = payload.winner_provider
    prize.winner_provider_user_id = payload.winner_provider_user_id
    prize.winning_number = payload.winning_number


def _apply_giveaway_payload(giveaway: Giveaway, payload: GiveawayWrite) -> None:
    giveaway.title = payload.title.strip()
    giveaway.description = (payload.description or '').strip() or None
    giveaway.is_active = payload.is_active
    giveaway.starts_at = payload.starts_at
    giveaway.ends_at = payload.ends_at
    giveaway.winners_count = payload.winners_count
    giveaway.telegram_community_url = (payload.telegram_community_url or "").strip() or None
    giveaway.telegram_chat_id = (payload.telegram_chat_id or "").strip() or None
    giveaway.telegram_reward_enabled = payload.telegram_reward_enabled
    giveaway.telegram_reward_numbers = payload.telegram_reward_numbers
    giveaway.vk_community_url = (payload.vk_community_url or "").strip() or None
    giveaway.vk_group_id = (payload.vk_group_id or "").strip() or None
    giveaway.vk_reward_enabled = payload.vk_reward_enabled
    giveaway.vk_reward_numbers = payload.vk_reward_numbers

    requested_prizes = _validate_giveaway_prizes(payload)
    existing_by_place = {prize.place_number: prize for prize in giveaway.prizes}
    requested_places = {prize.place_number for prize in requested_prizes}

    for existing_prize in list(giveaway.prizes):
        if existing_prize.place_number not in requested_places:
            giveaway.prizes.remove(existing_prize)

    for requested_prize in requested_prizes:
        existing_prize = existing_by_place.get(requested_prize.place_number)
        if existing_prize is None:
            existing_prize = GiveawayPrize()
            giveaway.prizes.append(existing_prize)
        _copy_giveaway_prize_fields(existing_prize, requested_prize)


@router.get("/giveaways", response_model=list[GiveawayRead])
def list_admin_giveaways(admin: AdminUser = Depends(require_admin), db: Session = Depends(get_db)) -> list[GiveawayRead]:
    _ = admin
    giveaways = db.execute(select(Giveaway).options(selectinload(Giveaway.prizes)).order_by(Giveaway.created_at.desc(), Giveaway.id.desc())).scalars().all()
    return [_giveaway_to_read(g) for g in giveaways]


@router.post("/giveaways", response_model=GiveawayRead, status_code=status.HTTP_201_CREATED)
def create_admin_giveaway(payload: GiveawayWrite, admin: AdminUser = Depends(require_admin), db: Session = Depends(get_db)) -> GiveawayRead:
    _ = admin
    giveaway = Giveaway(title=payload.title.strip() or "Розыгрыш", winners_count=payload.winners_count)
    _apply_giveaway_payload(giveaway, payload)
    db.add(giveaway)
    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Giveaway could not be saved because of conflicting data") from exc
    db.refresh(giveaway)
    giveaway = db.execute(select(Giveaway).options(selectinload(Giveaway.prizes)).where(Giveaway.id == giveaway.id)).scalar_one()
    return _giveaway_to_read(giveaway)


@router.put("/giveaways/{giveaway_id}", response_model=GiveawayRead)
def update_admin_giveaway(giveaway_id: int, payload: GiveawayWrite, admin: AdminUser = Depends(require_admin), db: Session = Depends(get_db)) -> GiveawayRead:
    _ = admin
    giveaway = db.execute(select(Giveaway).options(selectinload(Giveaway.prizes)).where(Giveaway.id == giveaway_id)).scalar_one_or_none()
    if giveaway is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Giveaway not found")
    _apply_giveaway_payload(giveaway, payload)
    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Giveaway could not be saved because of conflicting data") from exc
    giveaway = db.execute(select(Giveaway).options(selectinload(Giveaway.prizes)).where(Giveaway.id == giveaway_id)).scalar_one()
    return _giveaway_to_read(giveaway)


SOURCE_LABELS = {
    "subscription": "Основной номер за trial/подписку Bloom",
    "referral": "Номер за реферала",
    "telegram_subscription": "Номер за подписку на Telegram",
    "vk_subscription": "Номер за подписку на VK",
    "manual": "Ручное начисление администратором",
}

def _owner_name(client: ClientProfile | None) -> str | None:
    if client is None:
        return None
    full = client.full_name or " ".join(part for part in [client.telegram_first_name, client.telegram_last_name] if part).strip()
    return full or client.telegram_username or client.vk_username

def _giveaway_number_payload(number: GiveawayNumber, client: ClientProfile | None) -> dict[str, object]:
    user = client.user if client is not None else None
    return {
        "id": number.id, "number": number.number, "status": number.status, "is_active": is_number_active(number),
        "source": number.source, "source_label": SOURCE_LABELS.get(number.source, number.source),
        "created_at": number.created_at, "deactivated_at": number.deactivated_at, "deactivation_reason": number.deactivation_reason,
        "owner_name": _owner_name(client), "client_id": number.client_id,
        "telegram_id": client.telegram_user_id if client else None, "telegram_username": client.telegram_username if client else None,
        "vk_id": client.vk_user_id if client else None, "phone": user.phone if user else None, "email": client.contact_email or (user.email if user else None),
    }

@router.get("/giveaways/{giveaway_id}/entries")
def list_admin_giveaway_entries(
    giveaway_id: int, search_number: str | None = None, search_name: str | None = None, search_telegram: str | None = None, search_vk: str | None = None,
    source: str | None = None, active: bool | None = None, date_from: date | None = None, date_to: date | None = None,
    sort: str = "number", direction: str = "asc", admin: AdminUser = Depends(require_admin), db: Session = Depends(get_db)
) -> dict[str, object]:
    _ = admin
    if db.get(Giveaway, giveaway_id) is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Giveaway not found")
    stmt = select(GiveawayNumber, ClientProfile).join(ClientProfile, ClientProfile.id == GiveawayNumber.client_id, isouter=True).where(GiveawayNumber.giveaway_id == giveaway_id)
    if search_number: stmt = stmt.where(GiveawayNumber.number.ilike(f"%{search_number}%"))
    if source: stmt = stmt.where(GiveawayNumber.source == source)
    if active is not None: stmt = stmt.where(GiveawayNumber.is_active.is_(active), GiveawayNumber.status == ("active" if active else GiveawayNumber.status))
    if date_from: stmt = stmt.where(GiveawayNumber.created_at >= datetime.combine(date_from, datetime.min.time()))
    if date_to: stmt = stmt.where(GiveawayNumber.created_at < datetime.combine(date_to + timedelta(days=1), datetime.min.time()))
    rows = db.execute(stmt).all()
    items = [_giveaway_number_payload(n, c) for n, c in rows]
    if search_name: items = [i for i in items if search_name.lower() in str(i.get("owner_name") or "").lower()]
    if search_telegram: items = [i for i in items if search_telegram.lower() in (str(i.get("telegram_id") or "") + " " + str(i.get("telegram_username") or "")).lower()]
    if search_vk: items = [i for i in items if search_vk.lower() in str(i.get("vk_id") or "").lower()]
    reverse = direction == "desc"
    key_map = {"number": "number", "date": "created_at", "owner": "owner_name", "source": "source"}
    items.sort(key=lambda i: str(i.get(key_map.get(sort, "number")) or ""), reverse=reverse)
    summary = {"total_numbers": len(items), "active_numbers": sum(1 for i in items if i["is_active"]), "unique_participants": len({i["client_id"] for i in items}), "subscription_numbers": sum(1 for i in items if i["source"] == "subscription"), "referral_numbers": sum(1 for i in items if i["source"] == "referral"), "telegram_numbers": sum(1 for i in items if i["source"] == "telegram_subscription"), "vk_numbers": sum(1 for i in items if i["source"] == "vk_subscription")}
    return {"summary": summary, "items": items, "source_labels": SOURCE_LABELS}

@router.get("/giveaways/{giveaway_id}/entries/export.xlsx")
def export_admin_giveaway_entries(giveaway_id: int, admin: AdminUser = Depends(require_admin), db: Session = Depends(get_db)):
    _ = admin
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="openpyxl is required for XLSX export") from exc
    data = list_admin_giveaway_entries(giveaway_id, admin=admin, db=db)["items"]
    wb = Workbook(); ws = wb.active; ws.title = "Номера розыгрыша"
    headers = ["Номер", "Статус", "Источник", "ФИО", "Client ID", "Telegram ID", "Telegram username", "VK ID", "Телефон", "Email", "Дата начисления", "Дата деактивации", "Причина деактивации"]
    ws.append(headers)
    for cell in ws[1]: cell.font = Font(bold=True)
    for item in data:
        ws.append([item["number"], item["status"], item["source_label"], item["owner_name"], item["client_id"], item["telegram_id"], item["telegram_username"], item["vk_id"], item["phone"], item["email"], item["created_at"], item["deactivated_at"], item["deactivation_reason"]])
    ws.auto_filter.ref = ws.dimensions; ws.freeze_panes = "A2"
    widths = [14, 14, 34, 28, 12, 18, 22, 18, 18, 28, 22, 22, 32]
    for idx, width in enumerate(widths, 1): ws.column_dimensions[chr(64 + idx)].width = width
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1): row[0].number_format = "@"
    for row in ws.iter_rows(min_row=2, min_col=11, max_col=12):
        for cell in row: cell.number_format = "yyyy-mm-dd hh:mm:ss"
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    filename = f"bloom_giveaway_{giveaway_id}_{date.today().isoformat()}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})

@router.post("/giveaways/{giveaway_id}/social-subscriptions/recheck")
def recheck_admin_giveaway_social_subscriptions(giveaway_id: int, admin: AdminUser = Depends(require_admin), db: Session = Depends(get_db)) -> dict[str, int]:
    _ = admin
    giveaway = db.get(Giveaway, giveaway_id)
    if giveaway is None: raise HTTPException(status_code=404, detail="Giveaway not found")
    stats = recheck_giveaway_social_subscriptions(db, giveaway)
    db.commit()
    return stats


@router.get("/landing-settings", response_model=LandingSettingsRead)
def read_admin_landing_settings(
    admin: AdminUser = Depends(require_admin),
    db: Session = Depends(get_db),
) -> LandingSettingsRead:
    _ = admin
    return LandingSettingsRead.model_validate(build_admin_landing_settings_read(db))


@router.patch("/landing-settings", response_model=LandingSettingsRead)
def update_admin_landing_settings(
    payload: LandingSettingsUpdate,
    admin: AdminUser = Depends(require_admin),
    legacy_content_write: None = Depends(require_legacy_content_write_enabled),
    db: Session = Depends(get_db),
) -> LandingSettingsRead:
    _ = admin
    _ = legacy_content_write
    settings = get_or_create_landing_settings(db)
    update_data = payload.model_dump(exclude_unset=True)
    if update_data.get("partners_count_base") is not None and update_data.get("partners_count_display") is None:
        update_data["partners_count_display"] = update_data["partners_count_base"]
    if update_data.get("savings_total_base") is not None and update_data.get("savings_total") is None:
        update_data["savings_total"] = update_data["savings_total_base"]
    for field in ("members_count_base", "partners_count_display", "savings_total"):
        if field in update_data and update_data[field] is not None:
            setattr(settings, field, int(update_data[field]))
    for field in ("giveaway_title", "giveaway_current", "giveaway_subtitle", "giveaway_empty_text"):
        if field in update_data and update_data[field] is not None:
            setattr(settings, field, str(update_data[field]).strip())
    if "giveaway_items" in update_data and update_data["giveaway_items"] is not None:
        settings.giveaway_items = normalize_giveaway_items(update_data["giveaway_items"])
    db.add(settings)
    db.commit()
    db.refresh(settings)
    return LandingSettingsRead.model_validate(build_admin_landing_settings_read(db))


@router.get("/me", response_model=AdminUserRead)
def read_admin_me(admin: AdminUser = Depends(require_admin)) -> dict[str, object]:
    return {
        "id": admin.id,
        "email": admin.email,
        "role": admin.role,
        "legacy_content_write_enabled": settings.WEB_ADMIN_LEGACY_CONTENT_WRITE_ENABLED,
    }


@router.get("/activity", response_model=ActivityFeedRead)
def read_admin_activity(
    limit: int = 30,
    event_type: str | None = None,
    partner_id: int | None = None,
    admin: AdminUser = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ActivityFeedRead:
    _ = admin
    return build_admin_activity_feed(db, limit=limit, event_type=event_type, partner_id=partner_id)


@router.get("/verifications", response_model=list[AdminVerificationRead])
def list_admin_verifications(
    partner_id: int | None = None,
    client_id: int | None = None,
    status: str | None = None,
    admin: AdminUser = Depends(require_admin),
    db: Session = Depends(get_db),
) -> list[AdminVerificationRead]:
    _ = admin
    now = datetime.now(timezone.utc)
    normalize_expired_verifications(db, now=now, client_id=client_id, partner_id=partner_id)
    statement = (
        select(
            PrivilegeVerificationSession,
            ClientProfile.full_name.label("client_name"),
            Partner.name.label("partner_name"),
            City.id.label("city_id"),
            City.name.label("city_name"),
            PartnerOffer.title.label("offer_title"),
        )
        .join(ClientProfile, PrivilegeVerificationSession.client_id == ClientProfile.id)
        .join(Partner, PrivilegeVerificationSession.partner_id == Partner.id)
        .join(City, Partner.city_id == City.id)
        .outerjoin(PartnerOffer, PrivilegeVerificationSession.offer_id == PartnerOffer.id)
        .order_by(PrivilegeVerificationSession.created_at.desc(), PrivilegeVerificationSession.id.desc())
    )
    if partner_id is not None:
        statement = statement.where(PrivilegeVerificationSession.partner_id == partner_id)
    if client_id is not None:
        statement = statement.where(PrivilegeVerificationSession.client_id == client_id)
    statement = apply_verification_status_filter(statement, status, now=now)

    return [
        _admin_verification_to_read(session, client_name, partner_name, city_id, city_name, offer_title)
        for session, client_name, partner_name, city_id, city_name, offer_title in db.execute(statement).all()
    ]


@router.get("/payment-requests", response_model=list[AdminPaymentRequestRead])
def list_admin_payment_requests(
    status_filter: str | None = Query(default=None, alias="status"),
    client_id: int | None = None,
    limit: int = Query(default=50, ge=1, le=100),
    admin: AdminUser = Depends(require_admin),
    db: Session = Depends(get_db),
) -> list[AdminPaymentRequestRead]:
    _ = admin
    statement = (
        select(PaymentRequest)
        .options(
            selectinload(PaymentRequest.receipts),
            selectinload(PaymentRequest.client).selectinload(ClientProfile.user),
            selectinload(PaymentRequest.client).selectinload(ClientProfile.selected_city),
        )
        .order_by(PaymentRequest.created_at.desc(), PaymentRequest.id.desc())
        .limit(limit)
    )
    if status_filter is not None:
        statement = statement.where(PaymentRequest.status == status_filter)
    if client_id is not None:
        statement = statement.where(PaymentRequest.client_id == client_id)

    payment_requests = db.execute(statement).scalars().all()
    return [_admin_payment_request_to_read(payment_request) for payment_request in payment_requests]


@router.get("/payment-requests/{payment_request_id}", response_model=AdminPaymentRequestRead)
def read_admin_payment_request(
    payment_request_id: int,
    admin: AdminUser = Depends(require_admin),
    db: Session = Depends(get_db),
) -> AdminPaymentRequestRead:
    _ = admin
    payment_request = _get_admin_payment_request_or_404(db, payment_request_id)
    return _admin_payment_request_to_read(payment_request)


@router.post("/payment-requests/{payment_request_id}/approve", response_model=AdminPaymentRequestRead)
def approve_admin_payment_request(
    payment_request_id: int,
    payload: PaymentRequestApprove,
    admin: AdminUser = Depends(require_admin),
    db: Session = Depends(get_db),
) -> AdminPaymentRequestRead:
    payment_request = _get_admin_payment_request_or_404(db, payment_request_id)

    if payment_request.status == PaymentRequestStatus.approved.value:
        return _admin_payment_request_to_read(payment_request)
    if payment_request.status == PaymentRequestStatus.pending.value:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Payment request must be marked as paid before approval",
        )
    if payment_request.status == PaymentRequestStatus.rejected.value:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Rejected payment request cannot be approved")
    if payment_request.status != PaymentRequestStatus.paid.value:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unsupported payment request status")

    now = datetime.now(timezone.utc)
    latest_subscription = db.execute(
        select(Subscription)
        .where(Subscription.client_id == payment_request.client_id)
        .order_by(Subscription.ends_at.desc(), Subscription.id.desc())
        .limit(1)
    ).scalar_one_or_none()
    subscription_starts_at = now
    if (
        latest_subscription is not None
        and latest_subscription.status == SubscriptionStatus.active.value
        and as_aware_utc(latest_subscription.ends_at) > now
    ):
        subscription_starts_at = latest_subscription.ends_at

    subscription_ends_at = payload.access_until or subscription_starts_at + timedelta(days=payload.access_days or 30)
    if as_aware_utc(subscription_ends_at) <= as_aware_utc(subscription_starts_at):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Subscription end must be after start")

    payment_request.status = PaymentRequestStatus.approved.value
    payment_request.approved_at = now
    payment_request.rejected_at = None
    payment_request.admin_user_id = admin.id
    payment_request.updated_at = now
    payment_request.access_until = subscription_ends_at
    _append_admin_payment_request_comment(payment_request, payload.comment, prefix="Admin approval comment")

    db.add(
        Subscription(
            client_id=payment_request.client_id,
            status=SubscriptionStatus.active.value,
            starts_at=subscription_starts_at,
            ends_at=subscription_ends_at,
            source_payment_request_id=payment_request.id,
            source="paid",
        )
    )
    db.commit()
    payment_request = _get_admin_payment_request_or_404(db, payment_request.id)
    return _admin_payment_request_to_read(payment_request)


@router.post("/payment-requests/{payment_request_id}/reject", response_model=AdminPaymentRequestRead)
def reject_admin_payment_request(
    payment_request_id: int,
    payload: PaymentRequestReject,
    admin: AdminUser = Depends(require_admin),
    db: Session = Depends(get_db),
) -> AdminPaymentRequestRead:
    payment_request = _get_admin_payment_request_or_404(db, payment_request_id)

    if payment_request.status == PaymentRequestStatus.rejected.value:
        return _admin_payment_request_to_read(payment_request)
    if payment_request.status == PaymentRequestStatus.approved.value:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Approved payment request cannot be rejected")
    if payment_request.status not in {PaymentRequestStatus.pending.value, PaymentRequestStatus.paid.value}:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unsupported payment request status")

    now = datetime.now(timezone.utc)
    payment_request.status = PaymentRequestStatus.rejected.value
    payment_request.rejected_at = now
    payment_request.admin_user_id = admin.id
    payment_request.updated_at = now
    _append_admin_payment_request_comment(payment_request, payload.comment, prefix="Admin rejection comment")

    db.commit()
    payment_request = _get_admin_payment_request_or_404(db, payment_request.id)
    return _admin_payment_request_to_read(payment_request)


@router.get("/content-review", response_model=ContentReviewRead)
def read_admin_content_review(
    admin: AdminUser = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ContentReviewRead:
    _ = admin
    offer_rows = db.execute(
        select(PartnerOffer, Partner.name.label("partner_name"))
        .join(Partner, PartnerOffer.partner_id == Partner.id)
        .where(PartnerOffer.is_active.is_(False))
        .order_by(PartnerOffer.created_at.asc(), PartnerOffer.id.asc())
    ).all()
    photo_rows = db.execute(
        select(PartnerPhoto, Partner.name.label("partner_name"))
        .join(Partner, PartnerPhoto.partner_id == Partner.id)
        .where(PartnerPhoto.is_active.is_(False))
        .order_by(PartnerPhoto.created_at.asc(), PartnerPhoto.id.asc())
    ).all()

    return ContentReviewRead(
        offers=[
            ContentReviewOfferRead(
                id=offer.id,
                partner_id=offer.partner_id,
                partner_name=partner_name,
                title=offer.title,
                benefit_text=offer.benefit_text,
                description=offer.description,
                image_url=offer.image_url,
                created_at=offer.created_at,
            )
            for offer, partner_name in offer_rows
        ],
        photos=[
            ContentReviewPhotoRead(
                id=photo.id,
                partner_id=photo.partner_id,
                partner_name=partner_name,
                url=photo.url,
                alt_text=photo.alt_text,
                sort_order=photo.sort_order,
                created_at=photo.created_at,
            )
            for photo, partner_name in photo_rows
        ],
    )


@router.get("/users", response_model=list[AdminManagedUserRead])
def list_admin_users(
    role: str | None = None,
    is_active: bool | None = None,
    q: str | None = None,
    admin: AdminUser = Depends(require_admin),
    db: Session = Depends(get_db),
) -> list[User]:
    _ = admin
    statement = (
        select(
            User,
            ClientProfile.full_name.label("full_name"),
            ClientProfile.contact_email.label("contact_email"),
            ClientProfile.selected_city_id.label("selected_city_id"),
            City.name.label("selected_city_name"),
            ClientProfile.id.label("client_profile_id"),
            ClientProfile.vk_user_id.label("vk_user_id"),
            ClientProfile.vk_username.label("vk_username"),
            ClientProfile.telegram_user_id.label("telegram_user_id"),
            ClientProfile.telegram_username.label("telegram_username"),
            ClientProfile.trial_subscription_used_at.label("trial_subscription_used_at"),
        )
        .outerjoin(ClientProfile, ClientProfile.user_id == User.id)
        .outerjoin(City, City.id == ClientProfile.selected_city_id)
        .order_by(User.id.asc())
    )

    if role is not None:
        statement = statement.where(User.role == _normalize_user_role(role))
    if is_active is not None:
        statement = statement.where(User.is_active == is_active)
    if q is not None:
        search = q.strip()
        if search:
            pattern = f"%{search}%"
            statement = statement.where(or_(User.email.ilike(pattern), User.phone.ilike(pattern), ClientProfile.full_name.ilike(pattern), ClientProfile.contact_email.ilike(pattern), City.name.ilike(pattern), ClientProfile.vk_user_id.ilike(pattern), ClientProfile.vk_username.ilike(pattern), ClientProfile.telegram_user_id.ilike(pattern), ClientProfile.telegram_username.ilike(pattern)))

    rows = db.execute(statement).all()
    result: list[AdminManagedUserRead] = []
    now = datetime.now(timezone.utc)
    for (
        user,
        full_name,
        contact_email,
        selected_city_id,
        selected_city_name,
        client_profile_id,
        vk_user_id,
        vk_username,
        telegram_user_id,
        telegram_username,
        trial_subscription_used_at,
    ) in rows:
        normalized_email = (user.email or "").strip().lower()
        is_synthetic_email = normalized_email.startswith("vk_") and normalized_email.endswith("@vk.local") or normalized_email.endswith("@vk.local")
        vk_url = f"https://vk.com/{vk_username or ('id' + vk_user_id) }" if (vk_username or vk_user_id) else None
        telegram_url = f"https://t.me/{telegram_username}" if telegram_username else None
        active_subscription = None
        if client_profile_id is not None:
            active_subscription = db.execute(
                select(Subscription).where(
                    Subscription.client_id == client_profile_id,
                    Subscription.status == SubscriptionStatus.active.value,
                    Subscription.starts_at <= now,
                    Subscription.ends_at > now,
                ).order_by(Subscription.ends_at.desc(), Subscription.id.desc()).limit(1)
            ).scalar_one_or_none()
        active_subscription_type = "none"
        paid_subscription_status = "Не подключена"
        subscription_active_until = None
        if active_subscription is not None:
            subscription_active_until = active_subscription.ends_at
            active_subscription_type = "trial" if active_subscription.source == "trial" else "paid"
            if active_subscription_type == "paid":
                paid_subscription_status = "Подключена"
        display_name = (
            (full_name.strip() if isinstance(full_name, str) and full_name.strip() else None)
            or (contact_email.strip() if isinstance(contact_email, str) and contact_email.strip() else None)
            or (user.email.strip() if isinstance(user.email, str) and user.email.strip() else None)
            or f"Пользователь #{user.id}"
        )
        result.append(
            AdminManagedUserRead.model_validate(
                {
                    "id": user.id,
                    "email": user.email,
                    "phone": user.phone,
                    "role": user.role,
                    "is_active": user.is_active,
                    "full_name": full_name,
                    "contact_email": contact_email,
                    "selected_city_id": selected_city_id,
                    "selected_city_name": selected_city_name,
                    "vk_user_id": vk_user_id,
                    "vk_username": vk_username,
                    "vk_url": vk_url,
                    "telegram_user_id": telegram_user_id,
                    "telegram_username": telegram_username,
                    "telegram_url": telegram_url,
                    "trial_status": "Активировал" if trial_subscription_used_at is not None else "Не активировал",
                    "paid_subscription_status": paid_subscription_status,
                    "subscription_active_until": subscription_active_until,
                    "active_subscription_type": active_subscription_type,
                    "display_name": display_name,
                    "is_synthetic_email": is_synthetic_email,
                }
            )
        )
    return result


@router.post("/users", response_model=AdminManagedUserRead)
def create_admin_user(
    payload: AdminManagedUserCreate,
    admin: AdminUser = Depends(require_admin),
    db: Session = Depends(get_db),
) -> User:
    _ = admin
    email = _normalize_user_email(payload.email)
    phone = _normalize_user_phone(payload.phone)
    _ensure_user_contact_present(email, phone)
    role = _normalize_user_role(payload.role)
    password = _normalize_user_password(payload.password)
    _ensure_unique_user_identity(db, email=email, phone=phone)

    user = User(
        email=email,
        phone=phone,
        password_hash=hash_password(password),
        role=role,
        is_active=payload.is_active,
    )
    db.add(user)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise _user_duplicate_error() from None
    db.refresh(user)
    return user


@router.patch("/users/{user_id}", response_model=AdminManagedUserRead)
def update_admin_user(
    user_id: int,
    payload: AdminManagedUserUpdate,
    admin: AdminUser = Depends(require_admin),
    db: Session = Depends(get_db),
) -> User:
    _ = admin
    user = db.get(User, user_id)
    if user is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")

    update_data = payload.model_dump(exclude_unset=True)
    next_email = _normalize_user_email(update_data["email"]) if "email" in update_data else user.email
    next_phone = _normalize_user_phone(update_data["phone"]) if "phone" in update_data else user.phone
    _ensure_user_contact_present(next_email, next_phone)
    _ensure_unique_user_identity(db, email=next_email, phone=next_phone, exclude_user_id=user.id)

    if "email" in update_data:
        user.email = next_email
    if "phone" in update_data:
        user.phone = next_phone
    if "role" in update_data:
        user.role = _normalize_user_role(update_data["role"])
    if "is_active" in update_data:
        user.is_active = update_data["is_active"]
    if "password" in update_data:
        if update_data["password"] is not None:
            user.password_hash = hash_password(_normalize_user_password(update_data["password"]))

    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise _user_duplicate_error() from None
    db.refresh(user)
    return user


@router.delete("/users/{user_id}", response_model=AdminDeleteUserResponse)
def delete_admin_user(
    user_id: int,
    admin: AdminUser = Depends(require_admin),
    db: Session = Depends(get_db),
) -> dict[str, object]:
    return delete_user_with_relations(db=db, admin=admin, user_id=user_id)


@router.get("/cities", response_model=list[CityRead])
def list_admin_cities(
    admin: AdminUser = Depends(require_admin),
    db: Session = Depends(get_db),
) -> list[City]:
    _ = admin
    result = db.execute(select(City).order_by(City.sort_order.asc(), City.id.asc()))
    return list(result.scalars().all())


@router.post("/cities", response_model=CityRead)
def create_admin_city(
    payload: CityCreate,
    admin: AdminUser = Depends(require_admin),
    legacy_content_write: None = Depends(require_legacy_content_write_enabled),
    db: Session = Depends(get_db),
) -> City:
    _ = admin
    _ = legacy_content_write
    name = _strip_required(payload.name, "name")
    slug = _strip_required(payload.slug, "slug")
    _ensure_unique_city_identity(db, name=name, slug=slug)

    city = City(name=name, slug=slug, is_active=payload.is_active, sort_order=payload.sort_order)
    db.add(city)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise _city_duplicate_error() from None
    db.refresh(city)
    return city


@router.patch("/cities/{city_id}", response_model=CityRead)
def update_admin_city(
    city_id: int,
    payload: CityUpdate,
    admin: AdminUser = Depends(require_admin),
    legacy_content_write: None = Depends(require_legacy_content_write_enabled),
    db: Session = Depends(get_db),
) -> City:
    _ = admin
    _ = legacy_content_write
    city = db.get(City, city_id)
    if city is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="City not found")

    update_data = payload.model_dump(exclude_unset=True)
    next_name = _strip_required(update_data["name"], "name") if "name" in update_data else city.name
    next_slug = _strip_required(update_data["slug"], "slug") if "slug" in update_data else city.slug
    _ensure_unique_city_identity(db, name=next_name, slug=next_slug, exclude_city_id=city.id)

    for field, value in update_data.items():
        if field == "name":
            city.name = next_name
        elif field == "slug":
            city.slug = next_slug
        elif field == "is_active":
            city.is_active = value
        elif field == "sort_order":
            city.sort_order = value

    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise _city_duplicate_error() from None
    db.refresh(city)
    return city


@router.get("/categories", response_model=list[CategoryRead])
def list_admin_categories(
    admin: AdminUser = Depends(require_admin),
    db: Session = Depends(get_db),
) -> list[Category]:
    _ = admin
    result = db.execute(select(Category).order_by(Category.sort_order.asc(), Category.id.asc()))
    return list(result.scalars().all())


@router.post("/categories", response_model=CategoryRead)
def create_admin_category(
    payload: CategoryCreate,
    admin: AdminUser = Depends(require_admin),
    legacy_content_write: None = Depends(require_legacy_content_write_enabled),
    db: Session = Depends(get_db),
) -> Category:
    _ = admin
    _ = legacy_content_write
    name = _strip_category_required(payload.name, "name")
    slug = _strip_category_required(payload.slug, "slug")
    _ensure_unique_category_slug(db, slug=slug)

    category = Category(
        name=name,
        slug=slug,
        is_active=payload.is_active,
        sort_order=payload.sort_order if payload.sort_order is not None else 0,
    )
    db.add(category)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise _category_duplicate_error() from None
    db.refresh(category)
    return category


@router.patch("/categories/{category_id}", response_model=CategoryRead)
def update_admin_category(
    category_id: int,
    payload: CategoryUpdate,
    admin: AdminUser = Depends(require_admin),
    legacy_content_write: None = Depends(require_legacy_content_write_enabled),
    db: Session = Depends(get_db),
) -> Category:
    _ = admin
    _ = legacy_content_write
    category = db.get(Category, category_id)
    if category is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Category not found")

    update_data = payload.model_dump(exclude_unset=True)
    next_slug = _strip_category_required(update_data["slug"], "slug") if "slug" in update_data else category.slug
    if next_slug != category.slug:
        _ensure_unique_category_slug(db, slug=next_slug, exclude_category_id=category.id)

    for field, value in update_data.items():
        if field == "name":
            category.name = _strip_category_required(value, "name")
        elif field == "slug":
            category.slug = next_slug
        elif field == "is_active":
            category.is_active = value
        elif field == "sort_order":
            category.sort_order = value if value is not None else 0

    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise _category_duplicate_error() from None
    db.refresh(category)
    return category


@router.get("/partners", response_model=list[PartnerRead])
def list_admin_partners(
    city_id: int | None = None,
    is_active: bool | None = None,
    category_slug: str | None = None,
    q: str | None = None,
    admin: AdminUser = Depends(require_admin),
    db: Session = Depends(get_db),
) -> list[PartnerRead]:
    _ = admin
    statement = (
        select(Partner, City.name.label("city_name"), User.email.label("owner_email"))
        .join(City, Partner.city_id == City.id)
        .outerjoin(User, Partner.owner_user_id == User.id)
        .options(selectinload(Partner.categories))
        .order_by(Partner.sort_order.asc(), Partner.id.asc())
    )

    if city_id is not None:
        statement = statement.where(Partner.city_id == city_id)
    if is_active is not None:
        statement = statement.where(Partner.is_active == is_active)
    if category_slug is not None:
        normalized_category_slug = _normalize_category_slug(db, category_slug)
        statement = statement.where(
            or_(
                Partner.categories.any(Category.slug == normalized_category_slug),
                Partner.category_slug == normalized_category_slug,
            )
        )
    if q is not None:
        search = q.strip()
        if search:
            statement = statement.where(Partner.name.ilike(f"%{search}%"))

    return [
        _partner_to_read(partner, city_name, owner_email)
        for partner, city_name, owner_email in db.execute(statement).all()
    ]


@router.post("/partners", response_model=PartnerRead)
def create_admin_partner(
    payload: PartnerCreate,
    admin: AdminUser = Depends(require_admin),
    legacy_content_write: None = Depends(require_legacy_content_write_enabled),
    db: Session = Depends(get_db),
) -> PartnerRead:
    _ = admin
    _ = legacy_content_write
    _ensure_city_exists(db, payload.city_id)
    if payload.owner_user_id is not None:
        _get_partner_owner(db, payload.owner_user_id)

    category_ids = payload.category_ids
    partner = Partner(
        city_id=payload.city_id,
        owner_user_id=payload.owner_user_id,
        category_slug=_normalize_category_slug(db, payload.category_slug),
        name=_strip_partner_name(payload.name),
        is_active=payload.is_active,
        is_verified=payload.is_verified,
        sort_order=payload.sort_order,
    )
    for field in PARTNER_TEXT_FIELDS:
        setattr(partner, field, _normalize_optional_text(getattr(payload, field)))

    db.add(partner)
    if category_ids is not None:
        partner.categories = _get_categories_by_ids_or_400(db, category_ids)
        partner.category_slug = partner.categories[0].slug if partner.categories else None
    db.commit()
    db.refresh(partner)
    return _get_partner_read_or_404(db, partner.id)


@router.get("/partners/{partner_id}/analytics", response_model=PartnerAnalyticsRead)
def read_admin_partner_analytics(
    partner_id: int,
    admin: AdminUser = Depends(require_admin),
    db: Session = Depends(get_db),
) -> PartnerAnalyticsRead:
    _ = admin
    partner = _ensure_partner_exists(db, partner_id)
    return build_partner_analytics(db, partner)


@router.get("/partners/{partner_id}", response_model=PartnerRead)
def get_admin_partner(
    partner_id: int,
    admin: AdminUser = Depends(require_admin),
    db: Session = Depends(get_db),
) -> PartnerRead:
    _ = admin
    return _get_partner_read_or_404(db, partner_id)


@router.patch("/partners/{partner_id}", response_model=PartnerRead)
def update_admin_partner(
    partner_id: int,
    payload: PartnerUpdate,
    admin: AdminUser = Depends(require_admin),
    legacy_content_write: None = Depends(require_legacy_content_write_enabled),
    db: Session = Depends(get_db),
) -> PartnerRead:
    _ = admin
    _ = legacy_content_write
    partner = db.get(Partner, partner_id)
    if partner is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Partner not found")

    update_data = payload.model_dump(exclude_unset=True)
    if "city_id" in update_data:
        city_id = update_data["city_id"]
        if city_id is None:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="City not found")
        _ensure_city_exists(db, city_id)
        partner.city_id = city_id
    if "owner_user_id" in update_data:
        owner_user_id = update_data["owner_user_id"]
        if owner_user_id is not None:
            _get_partner_owner(db, owner_user_id)
        partner.owner_user_id = owner_user_id
    if "category_slug" in update_data:
        partner.category_slug = _normalize_category_slug(db, update_data["category_slug"])
    if "category_ids" in update_data:
        partner.categories = _get_categories_by_ids_or_400(db, update_data["category_ids"] or [])
        partner.category_slug = partner.categories[0].slug if partner.categories else None
    if "name" in update_data:
        partner.name = _strip_partner_name(update_data["name"])

    for field in PARTNER_TEXT_FIELDS:
        if field in update_data:
            setattr(partner, field, _normalize_optional_text(update_data[field]))
    for field in ("is_active", "is_verified", "sort_order"):
        if field in update_data:
            setattr(partner, field, update_data[field])

    db.commit()
    db.refresh(partner)
    return _get_partner_read_or_404(db, partner.id)


@router.post("/partners/{partner_id}/images")
async def upload_admin_partner_image(
    partner_id: int,
    kind: str,
    file: UploadFile = File(...),
    admin: AdminUser = Depends(require_admin),
    legacy_content_write: None = Depends(require_legacy_content_write_enabled),
    db: Session = Depends(get_db),
) -> dict[str, str]:
    _ = admin
    _ = legacy_content_write
    partner = _ensure_partner_exists(db, partner_id)
    normalized_kind = validate_image_kind(kind)
    image_url = await save_partner_image_upload(partner.id, normalized_kind, file)
    setattr(partner, f"{normalized_kind}_url", image_url)
    db.commit()
    return {"url": image_url, "kind": normalized_kind}


@router.post("/partners/{partner_id}/photos", response_model=PartnerPhotoUploadResponse)
async def upload_admin_partner_photo(
    partner_id: int,
    file: UploadFile = File(...),
    alt_text: str | None = Form(default=None),
    sort_order: int = Form(default=0),
    admin: AdminUser = Depends(require_admin),
    legacy_content_write: None = Depends(require_legacy_content_write_enabled),
    db: Session = Depends(get_db),
) -> PartnerPhoto:
    _ = admin
    _ = legacy_content_write
    partner = _ensure_partner_exists(db, partner_id)
    photo_url = await save_partner_photo_image_upload(partner.id, file)
    photo = PartnerPhoto(
        partner_id=partner.id,
        url=photo_url,
        alt_text=_normalize_optional_text(alt_text),
        sort_order=sort_order,
        is_active=True,
    )
    db.add(photo)
    db.commit()
    db.refresh(photo)
    return photo


@router.get("/partners/{partner_id}/photos", response_model=list[PartnerPhotoRead])
def list_admin_partner_photos(
    partner_id: int,
    admin: AdminUser = Depends(require_admin),
    db: Session = Depends(get_db),
) -> list[PartnerPhoto]:
    _ = admin
    partner = _ensure_partner_exists(db, partner_id)
    return list(
        db.execute(
            select(PartnerPhoto)
            .where(PartnerPhoto.partner_id == partner.id)
            .order_by(PartnerPhoto.sort_order.asc(), PartnerPhoto.created_at.asc())
        ).scalars().all()
    )


@router.patch("/partner-photos/{photo_id}", response_model=PartnerPhotoRead)
def update_admin_partner_photo(
    photo_id: int,
    payload: PartnerPhotoUpdate,
    admin: AdminUser = Depends(require_admin),
    legacy_content_write: None = Depends(require_legacy_content_write_enabled),
    db: Session = Depends(get_db),
) -> PartnerPhoto:
    _ = admin
    _ = legacy_content_write
    photo = db.get(PartnerPhoto, photo_id)
    if photo is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Partner photo not found")
    update_data = payload.model_dump(exclude_unset=True)
    if "alt_text" in update_data:
        photo.alt_text = _normalize_optional_text(update_data["alt_text"])
    for field in ("sort_order", "is_active"):
        if field in update_data:
            setattr(photo, field, update_data[field])
    db.commit()
    db.refresh(photo)
    return photo


@router.post("/partners/{partner_id}/qr-links", response_model=PartnerQrLinkRead)
def create_admin_partner_qr_link(
    partner_id: int,
    payload: PartnerQrLinkCreate,
    admin: AdminUser = Depends(require_admin),
    db: Session = Depends(get_db),
) -> PartnerQrLinkRead:
    _ = admin
    partner = _ensure_partner_exists(db, partner_id)
    slug = _normalize_or_generate_qr_slug(db, partner.id, payload.slug)

    qr_link = PartnerQrLink(
        partner_id=partner.id,
        slug=slug,
        deep_link_payload=_normalize_optional_text(payload.deep_link_payload),
        target_url=_normalize_optional_text(payload.target_url),
        is_active=payload.is_active,
    )
    db.add(qr_link)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise _qr_slug_duplicate_error() from None
    db.refresh(qr_link)
    return PartnerQrLinkRead.model_validate(qr_link_to_read(qr_link, partner_name=partner.name))


@router.get("/partners/{partner_id}/qr-links", response_model=list[PartnerQrLinkRead])
def list_admin_partner_qr_links(
    partner_id: int,
    admin: AdminUser = Depends(require_admin),
    db: Session = Depends(get_db),
) -> list[PartnerQrLinkRead]:
    _ = admin
    partner = _ensure_partner_exists(db, partner_id)
    links = db.execute(
        select(PartnerQrLink)
        .where(PartnerQrLink.partner_id == partner.id)
        .order_by(PartnerQrLink.id.asc())
    ).scalars().all()
    return [
        PartnerQrLinkRead.model_validate(qr_link_to_read(link, partner_name=partner.name))
        for link in links
    ]


@router.patch("/qr-links/{qr_link_id}", response_model=PartnerQrLinkRead)
def update_admin_qr_link(
    qr_link_id: int,
    payload: PartnerQrLinkUpdate,
    admin: AdminUser = Depends(require_admin),
    db: Session = Depends(get_db),
) -> PartnerQrLinkRead:
    _ = admin
    row = db.execute(
        select(PartnerQrLink, Partner.name.label("partner_name"))
        .join(Partner, PartnerQrLink.partner_id == Partner.id)
        .where(PartnerQrLink.id == qr_link_id)
    ).one_or_none()
    if row is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="QR link not found")
    qr_link, partner_name = row

    update_data = payload.model_dump(exclude_unset=True)
    if "slug" in update_data:
        qr_link.slug = _normalize_existing_qr_slug(db, update_data["slug"], exclude_qr_link_id=qr_link.id)
    if "deep_link_payload" in update_data:
        qr_link.deep_link_payload = _normalize_optional_text(update_data["deep_link_payload"])
    if "target_url" in update_data:
        qr_link.target_url = _normalize_optional_text(update_data["target_url"])
    if "is_active" in update_data:
        qr_link.is_active = update_data["is_active"]

    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise _qr_slug_duplicate_error() from None
    db.refresh(qr_link)
    return PartnerQrLinkRead.model_validate(qr_link_to_read(qr_link, partner_name=partner_name))


@router.get("/leads/partners", response_model=list[LeadStatsRead])
def list_admin_partner_lead_stats(
    admin: AdminUser = Depends(require_admin),
    db: Session = Depends(get_db),
) -> list[LeadStatsRead]:
    _ = admin
    rows = db.execute(
        select(
            Partner.id.label("partner_id"),
            Partner.name.label("partner_name"),
            City.id.label("city_id"),
            City.name.label("city_name"),
            PartnerQrLink.id.label("qr_link_id"),
            PartnerQrLink.slug.label("qr_slug"),
            func.count(LeadClick.id).label("total_clicks"),
        )
        .join(Partner, LeadClick.partner_id == Partner.id)
        .outerjoin(City, LeadClick.city_id == City.id)
        .outerjoin(PartnerQrLink, LeadClick.qr_link_id == PartnerQrLink.id)
        .group_by(Partner.id, Partner.name, City.id, City.name, PartnerQrLink.id, PartnerQrLink.slug)
        .order_by(func.count(LeadClick.id).desc(), Partner.id.asc())
    ).all()
    return [LeadStatsRead.model_validate(dict(row._mapping)) for row in rows]


@router.get("/partners/{partner_id}/offers", response_model=list[PartnerOfferRead])
def list_admin_partner_offers(
    partner_id: int,
    is_active: bool | None = None,
    admin: AdminUser = Depends(require_admin),
    db: Session = Depends(get_db),
) -> list[PartnerOfferRead]:
    _ = admin
    partner = _ensure_partner_exists(db, partner_id)
    statement = (
        select(PartnerOffer)
        .where(PartnerOffer.partner_id == partner.id)
        .order_by(PartnerOffer.sort_order.asc(), PartnerOffer.id.asc())
    )
    if is_active is not None:
        statement = statement.where(PartnerOffer.is_active == is_active)

    return [
        _partner_offer_to_read(offer, partner_name=partner.name)
        for offer in db.execute(statement).scalars().all()
    ]


@router.post("/partners/{partner_id}/offers", response_model=PartnerOfferRead)
def create_admin_partner_offer(
    partner_id: int,
    payload: PartnerOfferCreate,
    admin: AdminUser = Depends(require_admin),
    legacy_content_write: None = Depends(require_legacy_content_write_enabled),
    db: Session = Depends(get_db),
) -> PartnerOfferRead:
    _ = admin
    _ = legacy_content_write
    partner = _ensure_partner_exists(db, partner_id)
    _validate_offer_amounts(payload.base_price, payload.discount_percent)

    offer = PartnerOffer(
        partner_id=partner.id,
        title=_strip_offer_title(payload.title),
        base_price=payload.base_price,
        discount_percent=payload.discount_percent,
        is_active=payload.is_active,
        sort_order=payload.sort_order,
    )
    for field in PARTNER_OFFER_TEXT_FIELDS:
        setattr(offer, field, _normalize_optional_text(getattr(payload, field)))

    db.add(offer)
    db.commit()
    db.refresh(offer)
    return _partner_offer_to_read(offer, partner_name=partner.name)


@router.get("/offers/{offer_id}", response_model=PartnerOfferRead)
def get_admin_partner_offer(
    offer_id: int,
    admin: AdminUser = Depends(require_admin),
    db: Session = Depends(get_db),
) -> PartnerOfferRead:
    _ = admin
    return _get_partner_offer_read_or_404(db, offer_id)


@router.post("/offers/{offer_id}/image")
async def upload_admin_partner_offer_image(
    offer_id: int,
    file: UploadFile = File(...),
    admin: AdminUser = Depends(require_admin),
    legacy_content_write: None = Depends(require_legacy_content_write_enabled),
    db: Session = Depends(get_db),
) -> dict[str, str]:
    _ = admin
    _ = legacy_content_write
    offer = db.get(PartnerOffer, offer_id)
    if offer is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Offer not found")

    image_url = await save_partner_offer_image_upload(offer.partner_id, offer.id, file)
    offer.image_url = image_url
    db.commit()
    return {"url": image_url}


@router.patch("/offers/{offer_id}", response_model=PartnerOfferRead)
def update_admin_partner_offer(
    offer_id: int,
    payload: PartnerOfferUpdate,
    admin: AdminUser = Depends(require_admin),
    legacy_content_write: None = Depends(require_legacy_content_write_enabled),
    db: Session = Depends(get_db),
) -> PartnerOfferRead:
    _ = admin
    _ = legacy_content_write
    offer = db.get(PartnerOffer, offer_id)
    if offer is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Offer not found")

    update_data = payload.model_dump(exclude_unset=True)
    _validate_offer_amounts(update_data.get("base_price"), update_data.get("discount_percent"))

    if "title" in update_data:
        offer.title = _strip_offer_title(update_data["title"])
    for field in PARTNER_OFFER_TEXT_FIELDS:
        if field in update_data:
            setattr(offer, field, _normalize_optional_text(update_data[field]))
    for field in ("base_price", "discount_percent", "is_active", "sort_order"):
        if field in update_data:
            setattr(offer, field, update_data[field])

    db.commit()
    db.refresh(offer)
    return _get_partner_offer_read_or_404(db, offer.id)


def _admin_verification_to_read(
    session: PrivilegeVerificationSession,
    client_name: str | None,
    partner_name: str | None,
    city_id: int | None,
    city_name: str | None,
    offer_title: str | None,
) -> AdminVerificationRead:
    return AdminVerificationRead.model_validate(
        {
            "id": session.id,
            "client_id": session.client_id,
            "client_name": client_name,
            "partner_id": session.partner_id,
            "partner_name": partner_name,
            "city_id": city_id,
            "city_name": city_name,
            "offer_id": session.offer_id,
            "offer_title": offer_title,
            "code": session.code,
            "status": session.status,
            "source": session.source,
            "expires_at": session.expires_at,
            "confirmed_at": session.confirmed_at,
            "created_at": session.created_at,
            "ttl_seconds": ttl_seconds(session.expires_at),
        }
    )


def _get_admin_payment_request_or_404(db: Session, payment_request_id: int) -> PaymentRequest:
    payment_request = db.execute(
        select(PaymentRequest)
        .options(
            selectinload(PaymentRequest.receipts),
            selectinload(PaymentRequest.client).selectinload(ClientProfile.user),
            selectinload(PaymentRequest.client).selectinload(ClientProfile.selected_city),
        )
        .where(PaymentRequest.id == payment_request_id)
    ).scalar_one_or_none()
    if payment_request is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Payment request not found")
    return payment_request


def _admin_payment_request_to_read(payment_request: PaymentRequest) -> AdminPaymentRequestRead:
    client = payment_request.client
    user = client.user if client is not None else None
    city = client.selected_city if client is not None else None
    client_full_name = client.full_name if client is not None else None
    user_email = user.email if user is not None else None
    is_synthetic_email = bool(user_email and user_email.endswith("@vk.local"))
    vk_user_id = client.vk_user_id if client is not None else None
    vk_url = f"https://vk.com/id{vk_user_id}" if vk_user_id else None
    display_name = client_full_name or (client.contact_email if client is not None else None) or user_email
    if display_name is None and client is not None:
        display_name = f"Пользователь #{client.id}"
    return AdminPaymentRequestRead.model_validate(
        {
            "id": payment_request.id,
            "client_id": payment_request.client_id,
            "amount": payment_request.amount,
            "status": payment_request.status,
            "source": payment_request.source,
            "comment": payment_request.comment,
            "created_at": payment_request.created_at,
            "updated_at": payment_request.updated_at,
            "approved_at": payment_request.approved_at,
            "rejected_at": payment_request.rejected_at,
            "admin_user_id": payment_request.admin_user_id,
            "access_until": payment_request.access_until,
            "receipts": payment_request.receipts,
            "client_name": client_full_name,
            "client_full_name": client_full_name,
            "client_user_id": client.user_id if client is not None else None,
            "client_vk_user_id": vk_user_id,
            "user_id": client.user_id if client is not None else None,
            "user_email": user_email,
            "user_login": user_email,
            "user_phone": user.phone if user is not None else None,
            "full_name": client_full_name,
            "contact_email": client.contact_email if client is not None else None,
            "selected_city_name": city.name if city is not None else None,
            "vk_user_id": vk_user_id,
            "vk_url": vk_url,
            "display_name": display_name,
            "is_synthetic_email": is_synthetic_email,
        }
    )


def _append_admin_payment_request_comment(payment_request: PaymentRequest, comment: str | None, *, prefix: str) -> None:
    normalized_comment = _normalize_optional_text(comment)
    if normalized_comment is None:
        return
    comment_line = f"{prefix}: {normalized_comment}"
    if payment_request.comment:
        if comment_line in payment_request.comment:
            return
        payment_request.comment = f"{payment_request.comment}\n\n{comment_line}"
    else:
        payment_request.comment = comment_line


def _normalize_user_email(value: str | None) -> str | None:
    if value is None:
        return None
    normalized = value.strip().lower()
    return normalized or None


def _normalize_user_phone(value: str | None) -> str | None:
    if value is None:
        return None
    normalized = value.strip()
    return normalized or None


def _ensure_user_contact_present(email: str | None, phone: str | None) -> None:
    if email is None and phone is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Email or phone is required",
        )


def _normalize_user_role(value: str | None) -> str:
    normalized = value.strip().lower() if value is not None else ""
    if normalized not in ALLOWED_USER_ROLES:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid user role")
    return normalized


def _normalize_user_password(value: str) -> str:
    normalized = value.strip()
    if len(normalized) < 8:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Password must be at least 8 characters",
        )
    return normalized


def _ensure_unique_user_identity(
    db: Session,
    *,
    email: str | None,
    phone: str | None,
    exclude_user_id: int | None = None,
) -> None:
    conditions = []
    if email is not None:
        conditions.append(User.email == email)
    if phone is not None:
        conditions.append(User.phone == phone)
    if not conditions:
        return

    statement = select(User.id).where(or_(*conditions))
    if exclude_user_id is not None:
        statement = statement.where(User.id != exclude_user_id)
    duplicate_id = db.execute(statement.limit(1)).scalar_one_or_none()
    if duplicate_id is not None:
        raise _user_duplicate_error()


def _user_duplicate_error() -> HTTPException:
    return HTTPException(status_code=status.HTTP_409_CONFLICT, detail=USER_DUPLICATE_DETAIL)


def _qr_slug_duplicate_error() -> HTTPException:
    return HTTPException(status_code=status.HTTP_409_CONFLICT, detail="QR slug already exists")


def _invalid_qr_slug_error() -> HTTPException:
    return HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid QR slug")


def _ensure_unique_qr_slug(db: Session, slug: str, exclude_qr_link_id: int | None = None) -> None:
    statement = select(PartnerQrLink.id).where(PartnerQrLink.slug == slug)
    if exclude_qr_link_id is not None:
        statement = statement.where(PartnerQrLink.id != exclude_qr_link_id)
    duplicate_id = db.execute(statement.limit(1)).scalar_one_or_none()
    if duplicate_id is not None:
        raise _qr_slug_duplicate_error()


def _normalize_existing_qr_slug(
    db: Session,
    slug: str | None,
    *,
    exclude_qr_link_id: int | None = None,
) -> str:
    normalized = normalize_qr_slug(slug)
    if not is_valid_qr_slug(normalized):
        raise _invalid_qr_slug_error()
    assert normalized is not None
    _ensure_unique_qr_slug(db, normalized, exclude_qr_link_id=exclude_qr_link_id)
    return normalized


def _normalize_or_generate_qr_slug(db: Session, partner_id: int, slug: str | None) -> str:
    if slug is not None:
        return _normalize_existing_qr_slug(db, slug)
    for _ in range(5):
        generated = generate_qr_slug(partner_id)
        if is_valid_qr_slug(generated):
            existing_id = db.execute(
                select(PartnerQrLink.id).where(PartnerQrLink.slug == generated).limit(1)
            ).scalar_one_or_none()
            if existing_id is None:
                return generated
    raise _qr_slug_duplicate_error()


def _strip_required(value: str, field_name: str) -> str:
    normalized = value.strip()
    if not normalized:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"City {field_name} must not be empty",
        )
    return normalized


def _ensure_unique_city_identity(
    db: Session,
    *,
    name: str,
    slug: str,
    exclude_city_id: int | None = None,
) -> None:
    statement = select(City.id).where(or_(City.name == name, City.slug == slug))
    if exclude_city_id is not None:
        statement = statement.where(City.id != exclude_city_id)
    duplicate_id = db.execute(statement.limit(1)).scalar_one_or_none()
    if duplicate_id is not None:
        raise _city_duplicate_error()


def _city_duplicate_error() -> HTTPException:
    return HTTPException(status_code=status.HTTP_409_CONFLICT, detail=CITY_DUPLICATE_DETAIL)


def _ensure_city_exists(db: Session, city_id: int) -> City:
    city = db.get(City, city_id)
    if city is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="City not found")
    return city


def _get_partner_owner(db: Session, owner_user_id: int) -> User:
    owner = db.get(User, owner_user_id)
    if owner is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Owner user not found")
    if owner.role != UserRole.PARTNER.value:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Owner user must have partner role",
        )
    return owner


def _strip_partner_name(value: str | None) -> str:
    normalized = value.strip() if value is not None else ""
    if not normalized:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Partner name must not be empty",
        )
    return normalized


def _normalize_optional_text(value: str | None) -> str | None:
    if value is None:
        return None
    normalized = value.strip()
    return normalized or None


def _strip_category_required(value: str | None, field_name: str) -> str:
    normalized = value.strip() if value is not None else ""
    if not normalized:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Category {field_name} must not be empty",
        )
    return normalized


def _ensure_unique_category_slug(
    db: Session,
    *,
    slug: str,
    exclude_category_id: int | None = None,
) -> None:
    statement = select(Category.id).where(Category.slug == slug)
    if exclude_category_id is not None:
        statement = statement.where(Category.id != exclude_category_id)
    duplicate_id = db.execute(statement.limit(1)).scalar_one_or_none()
    if duplicate_id is not None:
        raise _category_duplicate_error()


def _category_duplicate_error() -> HTTPException:
    return HTTPException(status_code=status.HTTP_409_CONFLICT, detail=CATEGORY_DUPLICATE_DETAIL)


def _normalize_category_slug(db: Session, value: str | None) -> str | None:
    normalized = _normalize_optional_text(value)
    if normalized is None:
        return None
    category_count = db.execute(select(func.count()).select_from(Category)).scalar_one()
    if category_count == 0:
        if normalized not in WOMEN_CLUB_CATEGORY_SLUGS:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unknown category slug")
        return normalized

    category_id = db.execute(
        select(Category.id).where(Category.slug == normalized, Category.is_active.is_(True)).limit(1)
    ).scalar_one_or_none()
    if category_id is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unknown category slug")
    return normalized


def _ensure_partner_exists(db: Session, partner_id: int) -> Partner:
    partner = db.get(Partner, partner_id)
    if partner is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Partner not found")
    return partner


def _strip_offer_title(value: str | None) -> str:
    normalized = value.strip() if value is not None else ""
    if not normalized:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Offer title must not be empty",
        )
    return normalized


def _validate_offer_amounts(
    base_price: Decimal | None = None,
    discount_percent: Decimal | None = None,
) -> None:
    if base_price is not None and base_price < Decimal("0"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="base_price must be greater than or equal to 0",
        )
    if discount_percent is not None and (
        discount_percent < Decimal("0") or discount_percent > Decimal("100")
    ):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="discount_percent must be between 0 and 100",
        )


def _get_partner_offer_read_or_404(db: Session, offer_id: int) -> PartnerOfferRead:
    statement = (
        select(PartnerOffer, Partner.name.label("partner_name"))
        .join(Partner, PartnerOffer.partner_id == Partner.id)
        .where(PartnerOffer.id == offer_id)
    )
    row = db.execute(statement).one_or_none()
    if row is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Offer not found")
    offer, partner_name = row
    return _partner_offer_to_read(offer, partner_name=partner_name)


def _partner_offer_to_read(offer: PartnerOffer, partner_name: str | None) -> PartnerOfferRead:
    return PartnerOfferRead.model_validate(
        {
            "id": offer.id,
            "partner_id": offer.partner_id,
            "title": offer.title,
            "description": offer.description,
            "benefit_text": offer.benefit_text,
            "conditions": offer.conditions,
            "base_price": offer.base_price,
            "discount_percent": offer.discount_percent,
            "image_url": offer.image_url,
            "is_active": offer.is_active,
            "sort_order": offer.sort_order,
            "partner_name": partner_name,
        }
    )


def _get_partner_read_or_404(db: Session, partner_id: int) -> PartnerRead:
    statement = (
        select(Partner, City.name.label("city_name"), User.email.label("owner_email"))
        .join(City, Partner.city_id == City.id)
        .outerjoin(User, Partner.owner_user_id == User.id)
        .options(selectinload(Partner.categories))
        .where(Partner.id == partner_id)
    )
    row = db.execute(statement).one_or_none()
    if row is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Partner not found")
    partner, city_name, owner_email = row
    return _partner_to_read(partner, city_name, owner_email)


def _category_to_api_payload(category: Category | None) -> dict[str, object] | None:
    if category is None:
        return None
    name = getattr(category, "name", None) or getattr(category, "title", None) or getattr(category, "slug", None)
    title = getattr(category, "title", None) or name
    return {
        "id": getattr(category, "id", None),
        "title": title,
        "name": name,
        "slug": getattr(category, "slug", None),
        "sort_order": getattr(category, "sort_order", 0) or 0,
        "is_active": bool(getattr(category, "is_active", True)),
    }


def _partner_to_read(partner: Partner, city_name: str | None, owner_email: str | None) -> PartnerRead:
    categories = sorted(partner.categories, key=lambda c: (c.sort_order, c.name.lower(), c.id))
    first = categories[0] if categories else None
    legacy_slug = partner.category_slug
    first_category_payload = _category_to_api_payload(first)
    first_name = str(first_category_payload["name"]) if first_category_payload is not None else None
    first_slug = str(first_category_payload["slug"]) if first_category_payload is not None else None
    categories_payload = [_category_to_api_payload(category) for category in categories]
    normalized_categories_payload = [item for item in categories_payload if item is not None]
    return PartnerRead.model_validate(
        {
            "id": partner.id,
            "city_id": partner.city_id,
            "owner_user_id": partner.owner_user_id,
            "category_slug": first_slug or legacy_slug,
            "category_id": first.id if first is not None else None,
            "category_name": first_name,
            "category": first_category_payload,
            "categories": normalized_categories_payload,
            "category_ids": [c.id for c in categories],
            "category_slugs": [c.slug for c in categories],
            "name": partner.name,
            "description": partner.description,
            "address": partner.address,
            "phone": partner.phone,
            "website_url": partner.website_url,
            "social_url": partner.social_url,
            "instagram_url": partner.instagram_url,
            "vk_url": partner.vk_url,
            "telegram_url": partner.telegram_url,
            "whatsapp_url": partner.whatsapp_url,
            "map_url": partner.map_url,
            "working_hours": partner.working_hours,
            "logo_url": partner.logo_url,
            "cover_url": partner.cover_url,
            "is_active": partner.is_active,
            "is_verified": partner.is_verified,
            "sort_order": partner.sort_order,
            "city_name": city_name,
            "owner_email": owner_email,
        }
    )


def _get_categories_by_ids_or_400(db: Session, category_ids: list[int]) -> list[Category]:
    if not category_ids:
        return []
    categories = db.execute(select(Category).where(Category.id.in_(category_ids))).scalars().all()
    by_id = {category.id: category for category in categories}
    if len(by_id) != len(set(category_ids)):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Category not found")
    return [by_id[category_id] for category_id in dict.fromkeys(category_ids)]
